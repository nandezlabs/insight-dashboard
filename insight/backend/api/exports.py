"""
Export API endpoints for generating CSV, Excel, and PDF reports.
"""

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
import io
import csv
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER

from services.supabase_client import supabase

router = APIRouter(prefix="/api/exports", tags=["exports"])


@router.get("/csv")
async def export_csv(
    form_id: Optional[str] = Query(None, description="Filter by form ID"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
):
    """
    Export submissions as CSV file.
    Returns a downloadable CSV with all submission data.
    """
    try:
        # Build query
        query = supabase.table("submissions").select("*")
        
        if form_id:
            query = query.eq("form_id", form_id)
        
        if start_date:
            query = query.gte("created_at", f"{start_date}T00:00:00")
        
        if end_date:
            query = query.lte("created_at", f"{end_date}T23:59:59")
        
        # Fetch submissions
        response = query.order("created_at", desc=True).execute()
        submissions = response.data
        
        if not submissions:
            raise HTTPException(status_code=404, detail="No submissions found")
        
        # Fetch form templates for names
        form_ids = list(set(s["form_id"] for s in submissions))
        forms_response = supabase.table("form_templates").select("id,name").in_("id", form_ids).execute()
        forms_map = {f["id"]: f["name"] for f in forms_response.data}
        
        # Create CSV in memory
        output = io.StringIO()
        
        # Get all unique field keys across all submissions
        all_keys = set()
        for submission in submissions:
            if submission.get("data"):
                all_keys.update(submission["data"].keys())
        
        # Define CSV columns
        fieldnames = ["Submission ID", "Form Name", "Form Version", "Submitted At"] + sorted(all_keys)
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        # Write data rows
        for submission in submissions:
            row = {
                "Submission ID": submission["id"],
                "Form Name": forms_map.get(submission["form_id"], "Unknown"),
                "Form Version": submission.get("form_version", "N/A"),
                "Submitted At": submission["created_at"],
            }
            
            # Add form data fields
            if submission.get("data"):
                for key in all_keys:
                    value = submission["data"].get(key, "")
                    # Handle complex values
                    if isinstance(value, (list, dict)):
                        value = str(value)
                    row[key] = value
            
            writer.writerow(row)
        
        # Prepare response
        output.seek(0)
        filename = f"submissions_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/excel")
async def export_excel(
    form_id: Optional[str] = Query(None, description="Filter by form ID"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
):
    """
    Export submissions as Excel file with formatting.
    Returns a downloadable XLSX with styled sheets.
    """
    try:
        # Build query
        query = supabase.table("submissions").select("*")
        
        if form_id:
            query = query.eq("form_id", form_id)
        
        if start_date:
            query = query.gte("created_at", f"{start_date}T00:00:00")
        
        if end_date:
            query = query.lte("created_at", f"{end_date}T23:59:59")
        
        # Fetch submissions
        response = query.order("created_at", desc=True).execute()
        submissions = response.data
        
        if not submissions:
            raise HTTPException(status_code=404, detail="No submissions found")
        
        # Fetch form templates
        form_ids = list(set(s["form_id"] for s in submissions))
        forms_response = supabase.table("form_templates").select("id,name").in_("id", form_ids).execute()
        forms_map = {f["id"]: f["name"] for f in forms_response.data}
        
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Group submissions by form
        submissions_by_form = {}
        for submission in submissions:
            form_id_key = submission["form_id"]
            if form_id_key not in submissions_by_form:
                submissions_by_form[form_id_key] = []
            submissions_by_form[form_id_key].append(submission)
        
        # Create a sheet for each form
        for form_id_key, form_submissions in submissions_by_form.items():
            form_name = forms_map.get(form_id_key, "Unknown")
            # Excel sheet names have a 31-char limit
            sheet_name = form_name[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Get all unique field keys for this form
            all_keys = set()
            for submission in form_submissions:
                if submission.get("data"):
                    all_keys.update(submission["data"].keys())
            
            # Define headers
            headers = ["Submission ID", "Form Version", "Submitted At"] + sorted(all_keys)
            
            # Style header row
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data rows
            for row_idx, submission in enumerate(form_submissions, 2):
                ws.cell(row=row_idx, column=1, value=submission["id"][:8])
                ws.cell(row=row_idx, column=2, value=submission.get("form_version", "N/A"))
                ws.cell(row=row_idx, column=3, value=submission["created_at"])
                
                # Add form data fields
                if submission.get("data"):
                    for col_idx, key in enumerate(sorted(all_keys), 4):
                        value = submission["data"].get(key, "")
                        if isinstance(value, (list, dict)):
                            value = str(value)
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"submissions_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/pdf")
async def export_pdf(
    form_id: Optional[str] = Query(None, description="Filter by form ID"),
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
):
    """
    Export submissions as PDF report.
    Returns a downloadable PDF with formatted tables.
    """
    try:
        # Build query
        query = supabase.table("submissions").select("*")
        
        if form_id:
            query = query.eq("form_id", form_id)
        
        if start_date:
            query = query.gte("created_at", f"{start_date}T00:00:00")
        
        if end_date:
            query = query.lte("created_at", f"{end_date}T23:59:59")
        
        # Fetch submissions
        response = query.order("created_at", desc=True).execute()
        submissions = response.data
        
        if not submissions:
            raise HTTPException(status_code=404, detail="No submissions found")
        
        # Fetch form templates
        form_ids = list(set(s["form_id"] for s in submissions))
        forms_response = supabase.table("form_templates").select("id,name").in_("id", form_ids).execute()
        forms_map = {f["id"]: f["name"] for f in forms_response.data}
        
        # Create PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor("#1f2937"),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Title
        title = Paragraph("Submissions Export Report", title_style)
        elements.append(title)
        
        # Metadata
        metadata_style = styles["Normal"]
        metadata_text = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br/>"
        metadata_text += f"Total Submissions: {len(submissions)}"
        metadata = Paragraph(metadata_text, metadata_style)
        elements.append(metadata)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Group by form
        submissions_by_form = {}
        for submission in submissions:
            form_id_key = submission["form_id"]
            if form_id_key not in submissions_by_form:
                submissions_by_form[form_id_key] = []
            submissions_by_form[form_id_key].append(submission)
        
        # Create section for each form
        for form_id_key, form_submissions in submissions_by_form.items():
            form_name = forms_map.get(form_id_key, "Unknown")
            
            # Section header
            section_header = Paragraph(f"<b>{form_name}</b> ({len(form_submissions)} submissions)", styles["Heading2"])
            elements.append(section_header)
            elements.append(Spacer(1, 0.2 * inch))
            
            # Get field keys
            all_keys = set()
            for submission in form_submissions:
                if submission.get("data"):
                    all_keys.update(submission["data"].keys())
            
            field_keys = sorted(all_keys)[:5]  # Limit to 5 fields for PDF width
            
            # Create table data
            table_data = [["ID", "Date"] + field_keys]
            
            for submission in form_submissions:
                row = [
                    submission["id"][:8],
                    datetime.fromisoformat(submission["created_at"].replace("Z", "+00:00")).strftime("%m/%d/%Y")
                ]
                
                if submission.get("data"):
                    for key in field_keys:
                        value = submission["data"].get(key, "")
                        if isinstance(value, (list, dict)):
                            value = str(value)[:30]  # Truncate long values
                        else:
                            value = str(value)[:30]
                        row.append(value)
                else:
                    row.extend([""] * len(field_keys))
                
                table_data.append(row)
            
            # Create and style table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 0.4 * inch))
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        filename = f"submissions_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/summary")
async def export_summary():
    """
    Get export summary statistics.
    Returns total counts for forms and submissions available for export.
    """
    try:
        # Get forms count
        forms_response = supabase.table("form_templates").select("id", count="exact").execute()
        forms_count = forms_response.count or 0
        
        # Get submissions count
        submissions_response = supabase.table("submissions").select("id", count="exact").execute()
        submissions_count = submissions_response.count or 0
        
        # Get date range
        date_range = {"earliest": None, "latest": None}
        if submissions_count > 0:
            earliest = supabase.table("submissions").select("created_at").order("created_at", desc=False).limit(1).execute()
            latest = supabase.table("submissions").select("created_at").order("created_at", desc=True).limit(1).execute()
            
            if earliest.data:
                date_range["earliest"] = earliest.data[0]["created_at"]
            if latest.data:
                date_range["latest"] = latest.data[0]["created_at"]
        
        return {
            "status": "ready",
            "forms_count": forms_count,
            "submissions_count": submissions_count,
            "date_range": date_range,
            "available_formats": ["csv", "excel", "pdf"],
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get summary: {str(e)}")
